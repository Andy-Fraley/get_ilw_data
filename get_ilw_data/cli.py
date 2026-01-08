import typer
import locale
import datetime
import os
import glob
import subprocess
import re
import pandas as pd
from ansible_vault import Vault
from .config import Config
from .models import LoggingLevel
from .logging_utils import setup_logging
from .email_utils import send_admin_email
from .ccb_api import get_list_of_ilw_individuals, get_list_of_ilw_transactions
from .data_processing import (
    get_lists_from_file, write_lists_to_file, list_to_dataframe, preprocess_deceased_individuals,
    drop_or_remap_children_givers, merge_down_alternate_name, map_transaction_fam_ids, get_mapping_dicts,
    reload_names_and_emails, calculate_follow_ups, get_cell_loc, set_column_number_format, set_column_widths,
    get_pretty_emails_from_fam
)
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from copy import copy
import logging

app = typer.Typer()

TIMESTAMP_FORMAT = '%Y%m%d%H%M%S'

# COA abbreviation mapping
COA_ABBREV_MAP = {
    'P': 'Projects',
    'WF': 'Water Filters',
    'GD': 'General Donation',
    'A': 'Auctions',
    'ST': 'Sponsorships & Tickets'
}

# Reverse mapping: COA full name to abbreviation
COA_REVERSE_MAP = {v: k for k, v in COA_ABBREV_MAP.items()}

def parse_project_assignments(project_assignments_path):
    """
    Parse the Project Assignments tab and track recharacterizations.
    
    Returns:
        dict: Mapping of Find codes to total recharacterization amounts
    """
    recharacterizations = {}
    
    try:
        with pd.ExcelFile(project_assignments_path) as xlsx:
            df_project_assignments = pd.read_excel(xlsx, sheet_name='Project Assignments')
    except Exception as e:
        logging.error(f"Failed to read Project Assignments tab: {e}")
        raise
    
    # Pattern for validating the fixed-format suffix: Date-Dollar_Amount-COA_Abbreviation
    # We parse from the end backwards to allow dashes in First and Last names
    # Example: Rotary Club-Rich-Mar-20231119-$1,000.00-GD or Brooks-Jan-20250404-$5,635.52-GD
    # The last 3 components are always: Date (8 digits), Amount (with optional commas and cents), COA (letters)
    suffix_pattern = re.compile(r'^(.*?)-(\d{8})-\$?([\d,]+\.[\d]{2})-([A-Z]+)$')
    
    for index, row in df_project_assignments.iterrows():
        find_value = row.get('Find', '')
        match_value = row.get('Match', '')
        amount_raw = row.get('Amount', 0)
        placeholder_value = row.get('Placeholder Value', '')
        full_names = row.get('Full Name(s)', '')
        
        # Parse Amount column - handle currency formatting like "$2,743.13"
        if pd.isna(amount_raw):
            amount = 0.0
        elif isinstance(amount_raw, str):
            # Remove $ and commas, then convert to float
            amount = float(amount_raw.replace('$', '').replace(',', ''))
        else:
            amount = float(amount_raw)
        
        # Convert find_value to string and handle NaN/None
        if pd.isna(find_value):
            find_value = ''
        else:
            find_value = str(find_value).strip()
        
        # Convert match_value to string and handle NaN/None
        if pd.isna(match_value):
            match_value = ''
        else:
            match_value = str(match_value).strip()
        
        # Skip rows where Find=#N/A and Match contains NOT_RECEIVED in date portion
        # These are acknowledged as donations not yet received
        if find_value == '#N/A' and 'NOT_RECEIVED' in match_value:
            continue
        
        # Handle blank or #N/A entries (not yet matched)
        if find_value == '' or find_value == '#N/A':
            logging.warning(f"Funding is not yet matched for ${amount:,.2f} towards {placeholder_value} project by {full_names}")
            continue
        
        # Ignore *AUTO MATCH* entries
        if find_value == '*AUTO MATCH*':
            continue
        
        # Parse from the end: extract Date-Amount-COA, leaving Last-First at the beginning
        match = suffix_pattern.match(find_value)
        if not match:
            logging.error(f"Invalid Find column format: '{find_value}' does not follow expected pattern Last-First-Date-Dollar_Amount-COA_Abbreviation")
            continue
        
        name_part, date_str, dollar_amount_str, coa_abbrev = match.groups()
        
        # Now parse the name_part as Last-First (split on the last dash)
        # This allows dashes within Last or First names
        name_parts = name_part.rsplit('-', 1)
        if len(name_parts) != 2:
            logging.error(f"Invalid Find column format: '{find_value}' - cannot parse Last-First from '{name_part}'")
            continue
        
        last_name, first_name = name_parts
        
        # Validate COA abbreviation
        if coa_abbrev not in COA_ABBREV_MAP:
            logging.error(f"Invalid COA abbreviation in Find column: '{find_value}' has unknown COA abbreviation '{coa_abbrev}'. Valid values are: {', '.join(COA_ABBREV_MAP.keys())}")
            continue
        
        # Ignore 'P' (Projects) entries as they don't need recharacterization
        if coa_abbrev == 'P':
            continue
        
        # Parse the donation amount from the Find column
        donation_amount = float(dollar_amount_str.replace(',', ''))
        
        # Validate that recharacterization amount doesn't exceed donation amount
        if amount > donation_amount:
            logging.error(f"Recharacterization amount ${amount:,.2f} exceeds donation amount ${donation_amount:,.2f} in Find column: '{find_value}'")
            continue
        
        # This entry needs recharacterization
        matched_coa = COA_ABBREV_MAP[coa_abbrev]
        logging.debug(f"${amount:,.2f} of ${donation_amount:,.2f} contribution by {first_name} {last_name} needs to be recharacterized from {matched_coa} to Projects")
        
        # Add to recharacterization dictionary
        if find_value in recharacterizations:
            recharacterizations[find_value] += amount
        else:
            recharacterizations[find_value] = amount
    
    return recharacterizations

def create_match_string(row):
    """
    Create a Match string for a donation row in the format:
    Last-First-YYYYMMDD-$Amount-COA_Abbrev
    
    Args:
        row: A pandas Series representing a donation row
    
    Returns:
        str: Match string in the required format
    """
    last = str(row['Last'])
    first = str(row['First'])
    date_str = row['Date'].strftime('%Y%m%d')
    amount_str = f"${row['Amount']:,.2f}"
    coa_abbrev = COA_REVERSE_MAP.get(row['Simple COA'], '')
    
    if not coa_abbrev:
        return None
    
    return f"{last}-{first}-{date_str}-{amount_str}-{coa_abbrev}"

def apply_recharacterizations(df_donations, recharacterizations):
    """
    Apply recharacterizations to a donations DataFrame.
    
    For each donation row:
    1. Create a Match string
    2. Check if it exists in the recharacterizations dictionary
    3. If found:
       - If recharacterization amount equals donation amount: change COA to Projects
       - If recharacterization amount < donation amount: split into two rows
       - Add comments to the Comments field describing the recharacterization
    
    Args:
        df_donations: DataFrame of donations to recharacterize
        recharacterizations: Dictionary mapping Match strings to recharacterization amounts
    
    Returns:
        tuple: (DataFrame with recharacterizations applied, set of used match strings)
    """
    # Create a copy to avoid modifying the original
    df = df_donations.copy()
    
    # Ensure Comments column exists
    if 'Comments' not in df.columns:
        df['Comments'] = ''
    
    # Track which recharacterization entries are used
    used_entries = set()
    
    # List to collect new rows (for split donations)
    new_rows = []
    rows_to_drop = []
    
    for idx, row in df.iterrows():
        match_string = create_match_string(row)
        
        if match_string and match_string in recharacterizations:
            rechar_amount = recharacterizations[match_string]
            donation_amount = row['Amount']
            original_coa = row['Simple COA']
            date_str = row['Date'].strftime('%m/%d/%Y')
            
            # Mark this entry as used
            used_entries.add(match_string)
            
            if abs(rechar_amount - donation_amount) < 0.01:  # Equal (accounting for float precision)
                # Simple case: change COA to Projects
                df.at[idx, 'Simple COA'] = 'Projects'
                
                # Add comment to the recharacterized row
                comment = f"${donation_amount:,.2f} recharacterized from {original_coa} to Projects"
                existing_comment = row['Comments']
                if pd.isna(existing_comment) or existing_comment == '':
                    df.at[idx, 'Comments'] = comment
                else:
                    df.at[idx, 'Comments'] = f"{existing_comment}; {comment}"
                
                logging.debug(f"Recharacterized full amount ${donation_amount:,.2f} for {row['First']} {row['Last']} on {date_str} from {original_coa} to Projects")
            elif rechar_amount < donation_amount:
                # Split case: reduce original amount and create new Projects row
                remaining_amount = donation_amount - rechar_amount
                df.at[idx, 'Amount'] = remaining_amount
                
                # Add comment to the remaining portion (the part that stays in original COA)
                remaining_comment = f"${remaining_amount:,.2f} of ${donation_amount:,.2f} left as {original_coa}, and ${rechar_amount:,.2f} recharacterized from {original_coa} to Projects separately"
                existing_comment = row['Comments']
                if pd.isna(existing_comment) or existing_comment == '':
                    df.at[idx, 'Comments'] = remaining_comment
                else:
                    df.at[idx, 'Comments'] = f"{existing_comment}; {remaining_comment}"
                
                # Create new row for the recharacterized portion
                new_row = row.copy()
                new_row['Amount'] = rechar_amount
                new_row['Simple COA'] = 'Projects'
                
                # Add comment to the recharacterized portion (the part that goes to Projects)
                rechar_comment = f"${rechar_amount:,.2f} of ${donation_amount:,.2f} recharacterized from {original_coa} to Projects"
                new_row['Comments'] = rechar_comment
                new_rows.append((idx, new_row))
                
                logging.debug(f"Split donation for {row['First']} {row['Last']} on {date_str}: ${remaining_amount:,.2f} remains in {original_coa}, ${rechar_amount:,.2f} to Projects")
            else:
                # This shouldn't happen due to validation, but log if it does
                logging.error(f"Recharacterization amount ${rechar_amount:,.2f} exceeds donation amount ${donation_amount:,.2f} for match string: {match_string}")
    
    # Insert new rows right after their corresponding original rows
    # Sort by index in reverse order to maintain correct positions
    for orig_idx, new_row in sorted(new_rows, key=lambda x: x[0], reverse=True):
        # Get the position after the original row
        pos = df.index.get_loc(orig_idx) + 1
        # Insert the new row
        df = pd.concat([df.iloc[:pos], pd.DataFrame([new_row]), df.iloc[pos:]]).reset_index(drop=True)
    
    return df, used_entries

def check_inverse_recharacterizations(df_recharacterized_donations, df_original_donations, df_individuals, df_families, project_assignments_path):
    """
    Check for cases where Projects donations don't match Project Assignments.
    
    For years 2018+, compares each family's total Projects donations against their
    total Project Assignments amounts. Logs WARNING/DEBUG if donations exceed assignments,
    and ERROR if assignments exceed donations.
    
    Uses Match string from Project Assignments to find the actual donation in Original
    Donations, then uses that donation's Family ID (which has Override Fam ID already applied).
    
    Args:
        df_recharacterized_donations: DataFrame of recharacterized donations (for calculating Projects totals)
        df_original_donations: DataFrame of original donations (for Match string lookup to get Family ID)
        df_individuals: DataFrame of individuals (not used, kept for compatibility)
        df_families: DataFrame with family information (for Name(s) column)
        project_assignments_path: Path to project_assignments.xlsx file
    
    Returns:
        dict: Mapping of (year, family_id) to (donations_total, assignments_total) for cases needing inverse recharacterization
    """
    # Step 1: Calculate total Projects donations per family per year (2018+)
    # Use Recharacterized Donations to get the correct COA after recharacterizations
    family_total_projects_donations = {}
    for idx, row in df_recharacterized_donations.iterrows():
        if row['Simple COA'] != 'Projects':
            continue
        
        year = row['Year']
        if year < 2018:
            continue
        
        family_id = row['Family ID']
        
        if year not in family_total_projects_donations:
            family_total_projects_donations[year] = {}
        if family_id not in family_total_projects_donations[year]:
            family_total_projects_donations[year][family_id] = 0
        
        family_total_projects_donations[year][family_id] += row['Amount']
    
    # Step 2: Read Project Assignments
    try:
        with pd.ExcelFile(project_assignments_path) as xlsx:
            df_project_assignments = pd.read_excel(xlsx, sheet_name='Project Assignments')
    except Exception as e:
        logging.error(f"Failed to read Project Assignments tab: {e}")
        return {}
    
    # Step 3: Create a mapping of Match strings to donation rows in Original Donations
    # This allows us to find the actual Family ID (with Override Fam ID applied from MatchedTransactions)
    # Match strings in Project Assignments match the pre-recharacterized donations in Original Donations
    match_to_donation = {}
    for idx, row in df_original_donations.iterrows():
        match_string = create_match_string(row)
        if match_string:
            match_to_donation[match_string] = row
    
    # Step 4: Calculate total Project Assignments per family per year (2018+)
    # Use Find string to find the donation and get its Family ID
    family_total_project_assignments = {}
    
    for idx, row in df_project_assignments.iterrows():
        amount_raw = row.get('Amount', 0)
        find_value = row.get('Find', '')
        match_value = row.get('Match', '')
        
        # Parse Amount
        if pd.isna(amount_raw):
            amount = 0.0
        elif isinstance(amount_raw, str):
            amount = float(amount_raw.replace('$', '').replace(',', ''))
        else:
            amount = float(amount_raw)
        
        # Convert find_value to string and handle NaN/None
        if pd.isna(find_value):
            find_value = ''
        else:
            find_value = str(find_value).strip()
        
        # Convert match_value to string and handle NaN/None
        if pd.isna(match_value):
            match_value = ''
        else:
            match_value = str(match_value).strip()
        
        # Skip rows where Find=#N/A and Match contains NOT_RECEIVED in date portion
        if find_value == '#N/A' and 'NOT_RECEIVED' in match_value:
            continue
        
        # Skip if no find value or invalid find value
        if find_value == '' or find_value == '#N/A' or find_value == '*AUTO MATCH*':
            continue
        
        # Extract year from Find string
        suffix_pattern = re.compile(r'^(.*?)-(\d{8})-\$?([\d,]+\.[\d]{2})-([A-Z]+)$')
        match = suffix_pattern.match(find_value)
        if not match:
            logging.warning(f"Invalid Find format in Project Assignments: {find_value}")
            continue
        
        name_part, date_str, dollar_amount_str, coa_abbrev = match.groups()
        
        # Extract year
        if len(date_str) != 8:
            continue
        year = int(date_str[:4])
        
        if year < 2018:
            continue
        
        # Find the donation using the Find string to get the correct Family ID
        donation_row = match_to_donation.get(find_value)
        
        if donation_row is None:
            logging.warning(f"Could not find donation for Find string in Project Assignments: {find_value}")
            continue
        
        # Use the Family ID from the donation (which has Override Fam ID applied)
        family_id = donation_row['Family ID']
        
        # Add to totals
        if year not in family_total_project_assignments:
            family_total_project_assignments[year] = {}
        if family_id not in family_total_project_assignments[year]:
            family_total_project_assignments[year][family_id] = 0
        
        family_total_project_assignments[year][family_id] += amount
    
    # Step 5: Create family_id to Name(s) mapping from Families DataFrame
    family_names = {}
    for idx, row in df_families.iterrows():
        family_id = row.get('Family ID')
        name = row.get('Name(s)', '')
        if family_id:
            family_names[family_id] = name
    
    # Step 6: Compare totals and log warnings/errors
    # Also collect cases that need inverse recharacterization
    inverse_rechar_cases = {}  # (year, family_id) -> (donations_total, assignments_total)
    
    all_years = set(family_total_projects_donations.keys()) | set(family_total_project_assignments.keys())
    
    for year in sorted(all_years):
        donations_by_family = family_total_projects_donations.get(year, {})
        assignments_by_family = family_total_project_assignments.get(year, {})
        
        all_families = set(donations_by_family.keys()) | set(assignments_by_family.keys())
        
        for family_id in sorted(all_families):
            donations_total = donations_by_family.get(family_id, 0)
            assignments_total = assignments_by_family.get(family_id, 0)
            family_name = family_names.get(family_id, f'Family {family_id}')
            
            if abs(donations_total - assignments_total) <= 0.01:
                # Totals match - this is good
                continue
            elif donations_total > assignments_total:
                # Projects donations exceed project assignments - needs inverse recharacterization
                excess = donations_total - assignments_total
                
                # Store for inverse recharacterization
                inverse_rechar_cases[(year, family_id)] = (donations_total, assignments_total)
                
                # Log as DEBUG if $0 in assignments, WARNING otherwise
                if assignments_total == 0:
                    logging.debug(f"Family {family_id} ({family_name}) in {year}: Projects donations ${donations_total:,.2f} with $0 Project Assignments - will recharacterize to General Donation")
                else:
                    logging.warning(f"Family {family_id} ({family_name}) in {year}: Projects donations ${donations_total:,.2f} exceed Project Assignments ${assignments_total:,.2f} by ${excess:,.2f} - inverse recharacterization needed")
            else:
                # Project assignments exceed Projects donations - this is an error
                shortfall = assignments_total - donations_total
                logging.error(f"Family {family_id} ({family_name}) in {year}: Project Assignments ${assignments_total:,.2f} exceed Projects donations ${donations_total:,.2f} by ${shortfall:,.2f}")
    
    return inverse_rechar_cases

def apply_inverse_recharacterizations(df_donations, inverse_rechar_cases, df_families, project_assignments_path):
    """
    Propose inverse recharacterizations and add detailed comments to donations.
    
    For families with Projects donations exceeding Project Assignments:
    - If $0 in Project Assignments: recharacterize from Projects to General Donation (actual change)
    - If non-zero excess: propose specific inverse recharacterizations (comments only, no changes)
      - Case 1: Donation not in Project Assignments Find column -> propose 100% recharacterization
      - Case 2: Donation amount exceeds Project Assignment amount and overage equals family discrepancy -> propose split
    
    Args:
        df_donations: DataFrame of donations to modify
        inverse_rechar_cases: dict mapping (year, family_id) to (donations_total, assignments_total)
        df_families: DataFrame with family information (for Name(s) column)
        project_assignments_path: Path to project_assignments.xlsx file
    
    Returns:
        Modified DataFrame with inverse recharacterizations applied and comments added
    """
    # Create a copy to avoid modifying the original
    df = df_donations.copy()
    
    # Ensure Comments column exists
    if 'Comments' not in df.columns:
        df['Comments'] = ''
    
    # Create family_id to Name(s) mapping from Families DataFrame
    family_names = {}
    for idx, row in df_families.iterrows():
        family_id = row.get('Family ID')
        name = row.get('Name(s)', '')
        if family_id:
            family_names[family_id] = name
    
    # Read Project Assignments to get Find column values and amounts
    try:
        with pd.ExcelFile(project_assignments_path) as xlsx:
            df_project_assignments = pd.read_excel(xlsx, sheet_name='Project Assignments')
    except Exception as e:
        logging.error(f"Failed to read Project Assignments tab: {e}")
        return df
    
    # Build a set of Find values from Project Assignments (donations that ARE in Project Assignments)
    # Also track amounts for each Find value
    find_values_set = set()
    find_to_amount = {}  # Maps Find string to Project Assignment amount
    
    for idx, row in df_project_assignments.iterrows():
        find_value = row.get('Find', '')
        override_category = row.get('Override Category', '')
        amount_raw = row.get('Amount', 0)
        
        # Parse amount
        if pd.isna(amount_raw):
            amount = 0.0
        elif isinstance(amount_raw, str):
            amount = float(amount_raw.replace('$', '').replace(',', ''))
        else:
            amount = float(amount_raw)
        
        # Convert find_value to string and handle NaN/None
        if pd.isna(find_value):
            find_value = ''
        else:
            find_value = str(find_value).strip()
        
        # Skip invalid find values
        if find_value == '' or find_value == '#N/A' or find_value == '*AUTO MATCH*':
            continue
        
        # Add to set of Find values
        find_values_set.add(find_value)
        
        # Track amount for this Find value
        if find_value in find_to_amount:
            find_to_amount[find_value] += amount
        else:
            find_to_amount[find_value] = amount
    
    # Process each case that needs inverse recharacterization
    for (year, family_id), (donations_total, assignments_total) in inverse_rechar_cases.items():
        # Get family name
        family_name = family_names.get(family_id, f'Family {family_id}')
        excess = donations_total - assignments_total
        
        # Find all Projects donations for this family in this year
        mask = (df['Simple COA'] == 'Projects') & (df['Year'] == year) & (df['Family ID'] == family_id)
        
        if assignments_total == 0:
            # $0 in Project Assignments: recharacterize from Projects to General Donation (actual change)
            for idx in df[mask].index:
                # Recharacterize from Projects to General Donation
                df.at[idx, 'Simple COA'] = 'General Donation'
                
                # Add comment
                comment = f"Recharacterized from Projects to General Donation - no associated projects in Project Assignments for year {year} for {family_name}"
                existing_comment = df.at[idx, 'Comments']
                
                if pd.isna(existing_comment) or existing_comment == '':
                    df.at[idx, 'Comments'] = comment
                else:
                    df.at[idx, 'Comments'] = f"{existing_comment}; {comment}"
                
                logging.debug(f"Recharacterized ${df.at[idx, 'Amount']:,.2f} from Projects to General Donation for {family_name} (Family {family_id}) in {year}")
        else:
            # Non-zero excess: propose specific inverse recharacterizations
            # Check each Projects donation for this family/year
            for idx in df[mask].index:
                donation_row = df.loc[idx]
                match_string = create_match_string(donation_row)
                
                if not match_string:
                    continue
                
                donation_amount = donation_row['Amount']
                
                # Case 1: Donation NOT in Project Assignments Find column
                if match_string not in find_values_set:
                    comment = f"PROPOSE: Recharacterize 100% (${donation_amount:,.2f}) from Projects to General Donation - not referenced in Project Assignments Find column"
                    existing_comment = df.at[idx, 'Comments']
                    
                    if pd.isna(existing_comment) or existing_comment == '':
                        df.at[idx, 'Comments'] = comment
                    else:
                        df.at[idx, 'Comments'] = f"{existing_comment}; {comment}"
                    
                    logging.debug(f"Proposed 100% inverse recharacterization for ${donation_amount:,.2f} donation (not in Project Assignments) for {family_name} (Family {family_id}) in {year}")
                
                # Case 2: Donation IS in Project Assignments but amount exceeds Project Assignment amount
                elif match_string in find_to_amount:
                    project_assignment_amount = find_to_amount[match_string]
                    
                    if donation_amount > project_assignment_amount:
                        overage = donation_amount - project_assignment_amount
                        
                        # Check if overage equals the family's total excess (within tolerance)
                        if abs(overage - excess) <= 0.01:
                            comment = f"PROPOSE: Split this donation - leave ${project_assignment_amount:,.2f} as Projects (matching Project Assignment) and create new entry for ${overage:,.2f} as General Donation (overage)"
                            existing_comment = df.at[idx, 'Comments']
                            
                            if pd.isna(existing_comment) or existing_comment == '':
                                df.at[idx, 'Comments'] = comment
                            else:
                                df.at[idx, 'Comments'] = f"{existing_comment}; {comment}"
                            
                            logging.debug(f"Proposed split for ${donation_amount:,.2f} donation (keep ${project_assignment_amount:,.2f} as Projects, ${overage:,.2f} to General Donation) for {family_name} (Family {family_id}) in {year}")
    
    return df

@app.command()
def process(
    xlsx_input_file: str = typer.Option(None, help="Path of XLSX input file, which is normally Input.xlsx in the program directory."),
    xlsx_output_file: str = typer.Option(None, help="Path for XLSX output file. If not specified, defaults to ilw_data_[YYYYMMDDhhmmss].xlsx in the 'tmp' subdirectory."),
    use_file_cache: bool = typer.Option(False, help="Use file cache instead of pulling from CCB API."),
    no_email: bool = typer.Option(False, help="Do not send notification emails."),
    logging_level: str = typer.Option(LoggingLevel.warning.value, case_sensitive=False),
    before_after_csvs: bool = typer.Option(False, help="Create CSVs in before_after_csvs subdirectory capturing state before and after applying overlay and concatenation data."),
    get_now: bool = typer.Option(False, help="Force retrieval of fresh project_assignments.xlsx from remote source, ignoring 24-hour cache.")
):
    """
    Main processing pipeline for ILW data.
    """
    # Set up config and runtime state
    config = Config()
    locale.setlocale(locale.LC_ALL, '')
    config.datetime_start = datetime.datetime.now()
    config.curr_year = config.datetime_start.year
    config.datetime_start_string = config.datetime_start.strftime(TIMESTAMP_FORMAT)
    config.prog_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config.prog_name = os.path.basename(__file__)

    # Set up logging
    setup_logging(config, logging_level)

    # Retrieve project_assignments.xlsx before any data processing
    pull_script_path = '/Users/afraley/Documents/src/sh/pull_latest_ilw_data/pull_latest_project_assignments.sh'
    input_dir = os.path.join(config.prog_dir, 'input')
    os.makedirs(input_dir, exist_ok=True)
    
    pull_cmd = [pull_script_path, '--target-dir', input_dir]
    if get_now:
        pull_cmd.append('--get-now')
    
    try:
        result = subprocess.run(pull_cmd, check=True, capture_output=True, text=True)
        logging.info(f"Retrieved input/project_assignments.xlsx")
    except subprocess.CalledProcessError as e:
        logging.error(f"Failed to retrieve project_assignments.xlsx: {e.stderr}")
        raise RuntimeError(f"Failed to retrieve project_assignments.xlsx: {e.stderr}")
    except FileNotFoundError:
        logging.error(f"Pull script not found at {pull_script_path}")
        raise RuntimeError(f"Pull script not found at {pull_script_path}")

    # Parse project assignments and track recharacterizations
    project_assignments_path = os.path.join(input_dir, 'project_assignments.xlsx')
    recharacterizations = parse_project_assignments(project_assignments_path)

    # Pull info from vault file to support email notifications and enable CCB access
    secrets_files = glob.glob(os.path.join(config.prog_dir, '.secrets_*'))
    if len(secrets_files) > 1:
        raise RuntimeError(f"Found more than one .secrets_* file: {', '.join(secrets_files)}. Can only be one.")
    with open(secrets_files[0]) as file:
        vault_password = file.readline()
    vault_file_path = os.path.join(config.prog_dir, 'vault.yml')
    vault = Vault(vault_password)
    vault_data = vault.load(open(vault_file_path).read())
    if not no_email:
        if 'gmail' not in vault_data:
            raise RuntimeError(f"'gmail' is a required section in {vault_file_path} file")
        vault_data_gmail = vault_data['gmail']
        config.gmail_user = vault_data_gmail['user']
        config.gmail_password = vault_data_gmail['password']
        config.notification_target_email = vault_data_gmail['notify_target']
    vault_data_ccb = vault_data['ccb']
    config.ccb_app_username = vault_data_ccb['app_username']
    config.ccb_app_password = vault_data_ccb['app_password']
    config.ccb_subdomain = vault_data_ccb['subdomain']

    # Process command-line args
    if xlsx_input_file is None:
        try_xlsx_input_file = os.path.join(config.prog_dir, 'Input.xlsx')
        if os.path.exists(try_xlsx_input_file):
            xlsx_input_file = try_xlsx_input_file
        else:
            raise RuntimeError(f"Default input file, '{try_xlsx_input_file}', does not exist. Therefore, you must specify XLSX input file using --xlsx-input-file and this file must have 4 tabs: 'IndividualUpdate', 'IndividualConcat', 'CoaRemap', and 'ReachOutContacts'.")
    else:
        if not os.path.exists(xlsx_input_file):
            raise RuntimeError(f"Specified XLSX input file, '{xlsx_input_file}', not found.")
    
    # Validate and process xlsx_output_file
    if xlsx_output_file is not None:
        # Ensure it has .xlsx extension
        if not xlsx_output_file.lower().endswith('.xlsx'):
            raise RuntimeError(f"Output file '{xlsx_output_file}' must have a .xlsx extension.")
        
        # Convert to absolute path if relative
        # For relative paths, resolve them relative to the original working directory
        # that was active when the script was called, not the current working directory
        if not os.path.isabs(xlsx_output_file):
            # Get the original working directory from environment variable if set by shell script
            # Otherwise fall back to current working directory
            original_cwd = os.environ.get('ORIGINAL_CWD', os.getcwd())
            xlsx_output_file = os.path.join(original_cwd, xlsx_output_file)
        
        # Validate that the directory exists
        output_dir = os.path.dirname(xlsx_output_file)
        if not os.path.exists(output_dir):
            raise RuntimeError(f"Directory '{output_dir}' does not exist for output file '{xlsx_output_file}'.")
        
        # Check if directory is writable
        if not os.access(output_dir, os.W_OK):
            raise RuntimeError(f"Directory '{output_dir}' is not writable for output file '{xlsx_output_file}'.")

    # Data acquisition
    if use_file_cache:
        config_log = config.prog_dir  # for file cache
        typer.echo('Pulling transactions and individuals from file cache.')
        set_of_giving_family_ids, set_of_giving_individual_ids, list_of_ilw_transactions, list_of_ilw_individuals = get_lists_from_file(config_log)
        with pd.ExcelFile(xlsx_input_file) as xlsx:
            df_add_families = pd.read_excel(xlsx, sheet_name='NonGivingFamilies')
            set_of_giving_family_ids.update(df_add_families['Family ID'].tolist())
        typer.echo('Done pulling transactions and individuals from file cache.')
    else:
        typer.echo('Pulling transactions from CCB API...')
        set_of_giving_family_ids, set_of_giving_individual_ids, list_of_ilw_transactions = get_list_of_ilw_transactions(config)
        typer.echo('Done pulling transactions from CCB API.')
        typer.echo('Pulling individuals from CCB API...')
        with pd.ExcelFile(xlsx_input_file) as xlsx:
            df_add_families = pd.read_excel(xlsx, sheet_name='NonGivingFamilies')
            set_of_giving_family_ids.update(df_add_families['Family ID'].tolist())
        list_of_ilw_individuals = get_list_of_ilw_individuals(config, set_of_giving_family_ids)
        write_lists_to_file(config.prog_dir, set_of_giving_family_ids, set_of_giving_individual_ids, list_of_ilw_transactions, list_of_ilw_individuals)
        typer.echo('Done pulling individuals from CCB API.')

    # --- DataFrame merging and transformation ---
    list_of_individual_columns_to_drop = [
        'Limited Access User', 'Campus', 'Email Privacy Level',
        'General Communication', 'Mailing Area', 'Mailing Carrier Route', 'Mailing Privacy Level',
        'Home Phone Privacy Level', 'Work Phone Privacy Level', 'Mobile Phone Privacy Level', 'Fax',
        'Fax Phone Privacy Level', 'Pager Phone', 'Pager Phone Privacy Level', 'Emergency Phone',
        'Emergency Phone Privacy Level', 'Emergency Contact Name', 'Birthday Privacy Level',
        'Anniversary Privacy Level', 'Gender Privacy Level', 'Giving ID', 'Marital Status Privacy Level',
        'Home Area', 'Home Street', 'Home City', 'Home State', 'Home Zip', 'Home Country', 'Home Privacy Level',
        'Work Area', 'Work Street', 'Work City', 'Work State', 'Work Zip', 'Work Country', 'Work Privacy Level',
        'Other Area', 'Other Street', 'Other City', 'Other State', 'Other Zip', 'Other Country',
        'Other Privacy Level', 'School Name', 'School Grade', 'Family/Household Mailing Name',
        'Preferred Language', 'Ethnicity', 'Homebound Ministry', 'Allergies', 'Confirmed no allergies',
        'Allergies Privacy Level', 'Commitment Date', 'Commitment Story', 'Current Story', 'My Web Site',
        'Work Web Site', 'Military', 'Service(s) usually attended', 'Plugged In Privacy Level',
        'User Defined - Text 1', 'User Defined - Text 2', 'User Defined - Text 3', 'Pastr When Leav',
        'Pastr When Join', 'Transferred To', 'Transferred Frm', 'Baptized By', 'SK Indiv ID', 'Mailbox Number',
        'User Defined - Date 1', 'FBI Fingerprint', 'PA Criminal Chk', 'PA Child Abuse', 'Confirmed Date',
        'PA Sex Offender Registry', 'Mand Rpt Trng', 'Child/Youth Eml', 'Ethnicity', 'Photo Release', 'Confirmed',
        'Spirit Mailing', 'Custom Field Privacy Level', 'Personality Style', 'Spiritual Gifts', 'Passions',
        'Abilities', 'My Fit Privacy Level', 'Last logged in', 'Spiritual Maturity', 'Child Work Date Start',
        'Child Work Date Stop', 'Other ID', 'Sync ID']

    # --- Ensure df_ilw_transactions is initialized and modified as in get_ilw_data_ORIG.py ---
    df_ilw_transactions = list_to_dataframe(list_of_ilw_transactions)
    df_ilw_transactions['Date'] = pd.to_datetime(df_ilw_transactions['Date'], format='%Y-%m-%d')
    df_ilw_transactions['Transaction ID'] = df_ilw_transactions['Transaction ID'].astype(np.int64)
    df_ilw_transactions['Ind ID'] = df_ilw_transactions['Ind ID'].astype(np.int64)
    df_ilw_transactions['Family ID'] = df_ilw_transactions['Family ID'].astype(np.int64)
    df_ilw_transactions['Batch ID'] = df_ilw_transactions['Batch ID'].astype(np.int64)
    df_ilw_transactions['Amount'] = df_ilw_transactions['Amount'].astype(float)
    df_ilw_transactions.sort_values(by=['Date'], ascending=False, inplace=True)

    list_of_ilw_individuals = preprocess_deceased_individuals(list_of_ilw_individuals)
    df_ilw_individuals = list_to_dataframe(list_of_ilw_individuals, list_of_individual_columns_to_drop)
    df_ilw_individuals['Ind ID'] = df_ilw_individuals['Ind ID'].astype(np.int64)
    df_ilw_individuals['Family ID'] = df_ilw_individuals['Family ID'].astype(np.int64)
    df_ilw_transactions, df_ilw_individuals = drop_or_remap_children_givers(df_ilw_transactions,
        df_ilw_individuals, set_of_giving_individual_ids)
    df_ilw_individuals = merge_down_alternate_name(df_ilw_individuals)

    if before_after_csvs:
        df_ilw_individuals.to_csv(os.path.join(config.prog_dir, 'before_after_csvs', 'individuals_before.csv'),
                                  index=False)
        df_ilw_transactions.to_csv(os.path.join(config.prog_dir, 'before_after_csvs', 'transactions_before.csv'),
                                   index=False)
        logging.debug("Saved 'before' CSVs, prior to overlay and concat.")

    # Overlay and concat
    with pd.ExcelFile(xlsx_input_file) as xlsx:
        df_overlay = pd.read_excel(xlsx, sheet_name='IndividualUpdate')
        df_overlay['Ind ID'] = df_overlay['Ind ID'].astype(np.int64)
        df_overlay['Family ID'] = df_overlay['Family ID'].astype('Int64')
        df_ilw_transactions, set_of_giving_family_ids = map_transaction_fam_ids(df_ilw_transactions,
            df_ilw_individuals, df_overlay, set_of_giving_family_ids)
        df_overlay.set_index('Ind ID', inplace=True)
        df_ilw_individuals.set_index('Ind ID', inplace=True)
        df_ilw_individuals.update(df_overlay, overwrite=True)
        # Replace all <None> string values with NaN to make them blank in output
        df_ilw_individuals.replace('<None>', np.nan, inplace=True)
        df_ilw_individuals.reset_index(inplace=True)
        df_concat = pd.read_excel(xlsx, sheet_name='IndividualConcat').fillna('')
        df_concat['Ind ID'] = df_concat['Ind ID'].astype(np.int64)
        df_concat['Family ID'] = df_concat['Family ID'].astype(np.int64)
        df_ilw_individuals = pd.concat([df_ilw_individuals, df_concat])
        df_coa_remap = pd.read_excel(xlsx, sheet_name='CoaRemap')
        dict_coa_remap = {row['COA']: row['New COA'] for _, row in df_coa_remap.iterrows()}
        df_matched_transactions = pd.read_excel(xlsx, sheet_name='MatchedTransactions')
        df_ilw_transactions = pd.merge(df_ilw_transactions, df_matched_transactions,
            on='Transaction ID', how='left')
        df_ilw_transactions['Override Fam ID'] = df_ilw_transactions['Override Fam ID'].astype('Int64')
        df_ilw_transactions['Override COA Category'] = df_ilw_transactions['Override COA Category'].astype('string')
        df_ilw_transactions['Family ID'] = np.where(df_ilw_transactions['Override Fam ID'].isnull(),
            df_ilw_transactions['Family ID'], df_ilw_transactions['Override Fam ID'])
        df_ilw_transactions['COA Category'] = np.where(df_ilw_transactions['Override COA Category'].isnull(),
            df_ilw_transactions['COA Category'], df_ilw_transactions['Override COA Category'])
        df_ilw_transactions['Family ID'] = df_ilw_transactions['Family ID'].astype('Int64')
        df_ilw_transactions['COA Category'] = df_ilw_transactions['COA Category'].astype('string')
        df_ilw_transactions_w_new_family_id = df_ilw_transactions.loc[df_ilw_transactions['Family ID'] >= 100000]
        set_of_giving_family_ids.update(df_ilw_transactions_w_new_family_id['Family ID'].tolist())

    df_ilw_individuals.sort_values(by=['Last', 'First'], inplace=True)
    
    # Add Full Name column before Email column
    df_ilw_individuals['Full Name'] = df_ilw_individuals['First'] + ' ' + df_ilw_individuals['Last']
    
    # Add Full Email column after Email column
    df_ilw_individuals['Full Email'] = df_ilw_individuals.apply(
        lambda row: '' if (pd.isna(row['Email']) or row['Email'] == '' or 
                          str(row['First']).startswith('[DECEASED]'))
        else f"{row['Full Name']} <{row['Email']}>", axis=1
    )
    
    # Reorder columns to place Full Name before Email and Full Email after Email
    cols = df_ilw_individuals.columns.tolist()
    email_index = cols.index('Email')
    # Insert Full Name before Email
    cols.insert(email_index, cols.pop(cols.index('Full Name')))
    # Insert Full Email after Email (email_index + 1 because Full Name was inserted before)
    cols.insert(email_index + 2, cols.pop(cols.index('Full Email')))
    df_ilw_individuals = df_ilw_individuals[cols]
    
    # Check for duplicate First+Last name combinations
    name_groups = df_ilw_individuals.groupby(['First', 'Last'])['Ind ID'].apply(list)
    duplicates = name_groups[name_groups.apply(len) > 1]
    for (first, last), ind_ids in duplicates.items():
        ind_ids_str = ', '.join(str(id) for id in ind_ids)
        
        # Categorize by Individual ID ranges
        ccb_ids = [id for id in ind_ids if id < 100000]
        manual_ids = [id for id in ind_ids if id >= 100000]
        
        if len(manual_ids) == 0:
            # All IDs are from CCB (< 100000) - check mailing addresses
            # Get mailing addresses for all CCB IDs
            addresses = []
            for ind_id in ccb_ids:
                ind_row = df_ilw_individuals[df_ilw_individuals['Ind ID'] == ind_id].iloc[0]
                street = str(ind_row.get('Mailing Street', '')).strip()
                city = str(ind_row.get('Mailing City', '')).strip()
                state = str(ind_row.get('Mailing State', '')).strip()
                zip_code = str(ind_row.get('Mailing Zip', '')).strip()
                
                # Handle NaN values
                if street == 'nan': street = ''
                if city == 'nan': city = ''
                if state == 'nan': state = ''
                if zip_code == 'nan': zip_code = ''
                
                address = f"{street}, {city}, {state} {zip_code}".strip(', ')
                addresses.append(address)
            
            # Check if all addresses are the same
            unique_addresses = set(addresses)
            if len(unique_addresses) == 1:
                # Same address - likely true duplicate
                address = addresses[0] if addresses[0] else '(no address)'
                logging.error(f"Duplicate name found: '{first} {last}' - Individual IDs: {ind_ids_str} - May be duplicate individuals in CCB - Same mailing address: {address}")
            else:
                # Different addresses - likely different people with same name
                addresses_str = '; '.join([f"ID {ccb_ids[i]}: {addresses[i] if addresses[i] else '(no address)'}" for i in range(len(ccb_ids))])
                logging.debug(f"Duplicate name found: '{first} {last}' - Individual IDs: {ind_ids_str} - May be duplicate individuals in CCB - Different mailing addresses: {addresses_str}")
        elif len(ccb_ids) == 0:
            # All IDs are manual additions (>= 100000)
            logging.error(f"Duplicate name found: '{first} {last}' - Individual IDs: {ind_ids_str} - Two individuals added manually in IndividualConcat tab of Input.xlsx are likely duplicates of each other")
        else:
            # Mix of CCB and manual IDs
            logging.error(f"Duplicate name found: '{first} {last}' - Individual IDs: {ind_ids_str} - Individual manually added in IndividualConcat tab of Input.xlsx has likely been replaced by an individual added to CCB")

    if before_after_csvs:
        df_ilw_individuals.to_csv(os.path.join(config.prog_dir, 'before_after_csvs', 'individuals_after.csv'),
                                  index=False)
        df_ilw_transactions.to_csv(os.path.join(config.prog_dir, 'before_after_csvs', 'transactions_after.csv'),
                                   index=False)
        logging.debug("Saved 'after' CSVs, after overlay and concat.")

    # Family DataFrame
    list_families = [['Family ID', 'Name(s)', 'Full Email(s)', 'Primary ID', 'Spouse ID']]
    mapping_dicts = get_mapping_dicts(df_ilw_individuals)
    for family_id in set_of_giving_family_ids:
        couple_names, couple_emails, first_in_couple, second_in_couple = get_pretty_emails_from_fam(
            family_id, mapping_dicts, df_ilw_individuals)
        list_families.append([family_id, couple_names, couple_emails, first_in_couple, second_in_couple])
    df_ilw_families = list_to_dataframe(list_families)

    # DROP "Couple Email(s)" from "Individuals (CCB Overlaid"" tab
    #df_ilw_individuals = pd.merge(df_ilw_individuals, df_ilw_families[['Family ID', 'Email(s)']],
    #    left_on='Family ID', right_on='Family ID', how='left')
    #df_ilw_individuals.rename(columns={'Email(s)': 'Couple Email(s)'}, inplace=True)

    # !!! Add "Individual Email" here in "XYZ <abc@xyz.com>" format

    # Donations DataFrame
    df_ilw_donations = pd.merge(df_ilw_transactions, df_ilw_individuals, left_on='Ind ID', right_on='Ind ID', how='left')
    df_ilw_donations['COA Category'] = df_ilw_donations['COA Category'].map(dict_coa_remap)
    df_ilw_donations['Couple Emails'] = ''
    df_ilw_donations['Couple Names'] = ''
    reload_names_and_emails(df_ilw_donations, df_ilw_individuals)
    df_ilw_donations['Thank You Note'] = ''
    df_ilw_donations['Assigned Project'] = ''
    df_ilw_donations.rename(columns={'Family ID_x': 'Family ID', 'Mailing Zip_x': 'Mailing Zip',
        'Email': 'Donor Email', 'COA Category': 'Simple COA'}, inplace=True)
    df_ilw_donations = df_ilw_donations[['Date', 'Amount', 'First', 'Last', 'Thank You Note', 'Assigned Project',
        'Simple COA', 'Tax Deductible', 'Payment Type', 'Donor Email', 'Couple Emails', 'Couple Names',
        'Mailing Street', 'Mailing City', 'Mailing State', 'Mailing Zip', 'Home Phone', 'Mobile Phone', 'Ind ID',
        'Family ID', 'Comments']]
    calculate_follow_ups(df_ilw_donations)
    df_ilw_donations.sort_values(by=['Date'], ascending=False, inplace=True)
    df_ilw_donations['Year'] = pd.DatetimeIndex(df_ilw_donations['Date']).year

    # --- Excel Output and Formatting ---
    # Summarize donations by year to create Summary tab
    df_ilw_summary = df_ilw_donations.pivot_table(index=['Family ID'], columns=['Year'],
        values=['Amount'], aggfunc='sum')
    df_ilw_summary.columns = [str(c_list[1]) for c_list in df_ilw_summary.columns.values]
    empty_row = pd.Series([None] * len(df_ilw_summary.columns), dtype='float')
    for fam_id in df_add_families['Family ID'].values:
        df_ilw_summary.loc[fam_id] = empty_row
    df_ilw_sponsorships = df_ilw_donations[df_ilw_donations['Simple COA'] == 'Sponsorships & Tickets']. \
        pivot_table(index=['Family ID'], columns=['Year'], values=['Amount'], aggfunc='sum')
    df_ilw_sponsorships.columns = [str(c_list[1]) for c_list in df_ilw_sponsorships.columns.values]
    df_ilw_sponsorships['All-Time Sponsorships'] = df_ilw_sponsorships.select_dtypes(include=['number']).sum(axis=1)
    if config.datetime_start.month < 5 or (config.datetime_start.month == 5 and config.datetime_start.day < 8):
        last_year = config.datetime_start.year - 1
    else:
        last_year = config.datetime_start.year
    str_last_year = str(last_year)
    str_last_year_rename = str_last_year + ' Sponsorships'
    df_ilw_sponsorships = df_ilw_sponsorships[['All-Time Sponsorships', str_last_year]]
    df_ilw_sponsorships.rename(columns={str_last_year: str_last_year_rename}, inplace=True)
    df_ilw_summary = pd.merge(df_ilw_summary.reset_index(), df_ilw_sponsorships, on='Family ID', how='left')
    df_ilw_summary = pd.merge(df_ilw_summary.reset_index(),
        df_ilw_families, on='Family ID', how='left')

    # Add placeholder columns for Reach-Out XXXX sheet (optional, commented out)
    # df_ilw_summary['Reach-Out Completed'] = 'No'
    # df_ilw_summary['Notes / Status'] = ''
    # df_ilw_summary['Expected Sponsor Amount'] = ''
    # df_ilw_summary['Expected Other Amount'] = ''
    # df_ilw_summary['Paid Amount (updated by Andy)'] = ''
    # df_ilw_summary['Drop from Future Call Lists'] = 'No'

    df_ilw_summary['Lifetime Giving'] = ''
    df_ilw_summary['Last 5 Years Giving'] = ''
    df_ilw_summary = df_ilw_summary.merge(df_ilw_individuals, how='left',
        left_on='Primary ID', right_on='Ind ID')
    df_ilw_summary = df_ilw_summary.drop(columns=['Family ID_y'], axis=1)
    df_ilw_summary = df_ilw_summary.rename(columns={'Family ID_x': 'Family ID'})
    year_columns = ['Lifetime Giving', 'Last 5 Years Giving']
    for i in range(config.curr_year, 2012, -1):
        year_columns.append(str(i))
    df_ilw_summary.rename(columns={'Mobile Phone': 'Primary Mobile Phone'}, inplace=True)
    df_ilw_summary['Spouse Mobile Phone'] = df_ilw_summary.apply(
        lambda row: '' if pd.isnull(row['Spouse ID']) else mapping_dicts.ind2row[row['Spouse ID']]['Mobile Phone'],
        axis=1)
    
    # Filter year_columns to only include years that actually exist in the dataframe
    existing_year_columns = [col for col in year_columns if col in df_ilw_summary.columns]
    
    columns_list = ['Name(s)', 'All-Time Sponsorships', str_last_year_rename] + existing_year_columns + \
        ['Last', 'Full Email(s)', 'Home Phone', 'Primary Mobile Phone', 'Spouse Mobile Phone', 'Mailing Street', \
         'Mailing City', 'Mailing State', 'Mailing Zip', 'Family ID', 'Primary ID', 'Spouse ID']
    df_ilw_summary = df_ilw_summary[columns_list]

    # Add Simple COA column using dict_coa_remap
    df_ilw_transactions['Simple COA'] = df_ilw_transactions['COA Category'].map(dict_coa_remap)
    
    df_ilw_transactions = df_ilw_transactions[['Date', 'Amount', 'Name', 'Ind ID', 'Family ID',
        'Family Position', 'Gender', 'Age', 'Transaction ID', 'Batch ID', 'Batch Name',
        'Transaction Grouping', 'COA Category', 'Simple COA', 'Payment Type', 'Check Number', 'Memo', 'Tax Deductible',
        'Comments']]
    df_ilw_transactions.sort_values(by=['Date'], ascending=False, inplace=True)

    # Create memory-based summary of giving by family and year
    overall_giving = {}
    for index, row in df_ilw_donations.iterrows():
        if not row['Family ID'] in overall_giving:
            overall_giving[row['Family ID']] = {}
        if not row['Year'] in overall_giving[row['Family ID']]:
            overall_giving[row['Family ID']][row['Year']] = {}
        if not row['Simple COA'] in overall_giving[row['Family ID']][row['Year']]:
            overall_giving[row['Family ID']][row['Year']][row['Simple COA']] = 0
        overall_giving[row['Family ID']][row['Year']][row['Simple COA']] += row['Amount']

    # Create Recharacterized Donations tab by applying recharacterizations
    # Need to do this before dropping Year column since apply_recharacterizations needs the data
    df_ilw_recharacterized, used_rechar_entries = apply_recharacterizations(df_ilw_donations, recharacterizations)
    
    # Validate: Check for unused recharacterization entries
    unused_entries = set(recharacterizations.keys()) - used_rechar_entries
    if unused_entries:
        logging.error(f"Found {len(unused_entries)} recharacterization entries that were not matched:")
        for entry in sorted(unused_entries):
            logging.error(f"  Unmatched recharacterization entry: {entry} (${recharacterizations[entry]:,.2f})")
    
    # Validate: Total amounts should be equal between Original and Recharacterized Donations
    original_total = df_ilw_donations['Amount'].sum()
    recharacterized_total = df_ilw_recharacterized['Amount'].sum()
    if abs(original_total - recharacterized_total) > 0.01:  # Allow for small floating point differences
        logging.error(f"Total amounts do not match: Original Donations = ${original_total:,.2f}, Recharacterized Donations = ${recharacterized_total:,.2f}, Difference = ${abs(original_total - recharacterized_total):,.2f}")
    else:
        logging.debug(f"Total amounts validated: Original = ${original_total:,.2f}, Recharacterized = ${recharacterized_total:,.2f}")
    
    # Check for inverse recharacterization cases (for years 2018+)
    # This validates that Projects donations match Project Assignments per family/year
    # Uses Original Donations for Match string lookup to get correct Family ID (with Override Fam ID applied)
    inverse_rechar_cases = check_inverse_recharacterizations(df_ilw_recharacterized, df_ilw_donations, df_ilw_individuals, df_ilw_families, project_assignments_path)
    
    # Apply inverse recharacterizations and add comments
    # - $0 Project Assignments: recharacterize Projects to General Donation
    # - Non-zero excess: propose specific inverse recharacterizations (comments only)
    df_ilw_recharacterized = apply_inverse_recharacterizations(df_ilw_recharacterized, inverse_rechar_cases, df_ilw_families, project_assignments_path)
    
    # Validate total amounts after inverse recharacterizations
    recharacterized_total_after_inverse = df_ilw_recharacterized['Amount'].sum()
    if abs(original_total - recharacterized_total_after_inverse) > 0.01:
        logging.error(f"Total amounts do not match after inverse recharacterizations: Original Donations = ${original_total:,.2f}, Recharacterized Donations = ${recharacterized_total_after_inverse:,.2f}, Difference = ${abs(original_total - recharacterized_total_after_inverse):,.2f}")
    else:
        logging.debug(f"Total amounts validated after inverse recharacterizations: Original = ${original_total:,.2f}, Recharacterized = ${recharacterized_total_after_inverse:,.2f}")
    
    # Drop Year column from both Donations tabs
    df_ilw_donations = df_ilw_donations.drop(columns=['Year'], axis=1)
    df_ilw_recharacterized = df_ilw_recharacterized.drop(columns=['Year'], axis=1)

    # Write to output file
    if xlsx_output_file is not None:
        output_filename = xlsx_output_file
    else:
        # Ensure tmp directory exists
        tmp_dir = os.path.join(config.prog_dir, 'tmp')
        os.makedirs(tmp_dir, exist_ok=True)
        output_filename = os.path.join(tmp_dir, f'ilw_data_{config.datetime_start_string}.xlsx')
    with pd.ExcelWriter(output_filename) as writer:
        df_ilw_summary.to_excel(writer, sheet_name='Summary By Year', index=False)
        df_ilw_donations.to_excel(writer, sheet_name='Original Donations', index=False)
        df_ilw_recharacterized.to_excel(writer, sheet_name='Recharacterized Donations', index=False)
        df_ilw_individuals.to_excel(writer, sheet_name='Individuals', index=False)
        df_ilw_transactions.to_excel(writer, sheet_name='Transactions', index=False)
        df_ilw_families.to_excel(writer, sheet_name='Families', index=False)

    # Reload workbook for formatting
    workbook = openpyxl.load_workbook(output_filename)

    # Original Donations sheet formatting
    worksheet = workbook['Original Donations']
    filters = worksheet.auto_filter
    filters.ref = 'A1:U' + str(worksheet.max_row)
    set_column_number_format(worksheet, 'A', 'm/d/yy')
    set_column_number_format(worksheet, 'B', '$#,##0.00')
    column_widths = {
        'A': 10, 'B': 13, 'C': 18, 'D': 19, 'E': 18, 'F': 19, 'G': 19, 'H': 17, 'I': 17, 'J': 45,
        'K': 110, 'L': 30, 'M': 69, 'N': 19, 'O': 16, 'P': 15, 'Q': 22, 'R': 22, 'S': 10, 'T': 14, 'U': 39
    }
    set_column_widths(worksheet, column_widths)

    # Recharacterized Donations sheet formatting
    worksheet = workbook['Recharacterized Donations']
    filters = worksheet.auto_filter
    filters.ref = 'A1:U' + str(worksheet.max_row)
    set_column_number_format(worksheet, 'A', 'm/d/yy')
    set_column_number_format(worksheet, 'B', '$#,##0.00')
    column_widths = {
        'A': 10, 'B': 13, 'C': 18, 'D': 19, 'E': 18, 'F': 19, 'G': 19, 'H': 17, 'I': 17, 'J': 45,
        'K': 110, 'L': 30, 'M': 69, 'N': 19, 'O': 16, 'P': 15, 'Q': 22, 'R': 22, 'S': 10, 'T': 14, 'U': 39
    }
    set_column_widths(worksheet, column_widths)

    # Summary sheet formatting
    worksheet = workbook['Summary By Year']
    filters = worksheet.auto_filter
    filters.ref = 'A1:AD' + str(worksheet.max_row)
    for c in range(ord('B'), ord('B') + len(year_columns)):
        set_column_number_format(worksheet, chr(c), '$#,##0.00')
    column_widths = {
        'A': 37, 'B': 16, 'C': 14, 'D': 12, 'E': 13, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10,
        'L': 10, 'M': 10, 'N': 10, 'O': 10, 'P': 10, 'Q': 10, 'R': 10, 'S': 19, 'T': 108, 'U': 22, 'V': 24,
        'W': 23, 'X': 70, 'Y': 19, 'Z': 17, 'AA': 15, 'AB': 13, 'AC': 14, 'AD': 14
    }
    set_column_widths(worksheet, column_widths)
    # Insert "Lifetime" (column) giving SUM() formula
    for i in range(2, worksheet.max_row+1):
        worksheet['D' + str(i)] = '=SUM(F' + str(i) + ':' + chr(67 + len(year_columns)) + str(i) + ')'
    # Insert "Last 5 Years" (column) giving SUM() formula
    for i in range(2, worksheet.max_row+1):
        worksheet['E' + str(i)] = '=SUM(F' + str(i) + ':' + 'K' + str(i) + ')'

    # Insert giving data as comments in Summary tab
    for fam_id in overall_giving:
        for year in overall_giving[fam_id]:
            comment_string = ''
            for coa_cat in overall_giving[fam_id][year]:
                if comment_string != '':
                    comment_string += '\n'
                comment_string += f'{coa_cat}: {overall_giving[fam_id][year][coa_cat]}'
            # get_cell_loc expects curr_year, so pass config.curr_year
            col_num, row_num = get_cell_loc(df_ilw_summary, fam_id, year, config.curr_year)
            worksheet.cell(row=row_num, column=col_num).comment =  Comment(comment_string, config.prog_name)

    # Wrap column headers for columns B-E
    for col in ['B', 'C', 'D', 'E']:
        new_align = copy(worksheet[col+'1'].alignment)
        new_align.wrap_text = True
        worksheet[col+'1'].alignment = new_align
    # Align header column bottom
    for cell in worksheet["1:1"]:
        new_align = copy(cell.alignment)
        new_align.vertical = 'bottom'
        cell.alignment = new_align

    # Individuals sheet formatting
    worksheet = workbook['Individuals']
    filters = worksheet.auto_filter
    filters.ref = 'A1:AQ' + str(worksheet.max_row)
    column_widths = {
        'A': 11, 'B': 14, 'C': 18, 'D': 11, 'E': 17, 'F': 19, 'G': 17, 'H': 20, 'I': 11, 'J': 33, 'K': 30, 'L': 47, 
        'M': 70, 'N': 20, 'O': 17, 'P': 15, 'Q': 19, 'R': 19, 'S': 21, 'T': 22, 'U': 16, 'V': 22, 'W': 18, 'X': 13, 
        'Y': 16, 'Z': 13, 'AA': 17, 'AB': 28, 'AC': 18, 'AD': 17, 'AE': 22, 'AF': 23, 'AG': 17, 'AH': 18, 'AI': 11, 
        'AJ': 13, 'AK': 25, 'AL': 25, 'AM': 21, 'AN': 13, 'AO': 18, 'AP': 19, 'AQ': 24
    }
    set_column_widths(worksheet, column_widths)
    
    # Color the Full Name and Full Email column headers light blue
    from openpyxl.styles import PatternFill
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    worksheet['J1'].fill = light_blue_fill  # Full Name column
    worksheet['L1'].fill = light_blue_fill  # Full Email column

    # Transactions sheet formatting
    worksheet = workbook['Transactions']
    filters = worksheet.auto_filter
    filters.ref = 'A1:S' + str(worksheet.max_row)
    set_column_number_format(worksheet, 'A', 'm/d/yy')
    set_column_number_format(worksheet, 'B', '$#,##0.00')
    column_widths = {
        'A': 10, 'B': 13, 'C': 33, 'D': 11, 'E': 14, 'F': 18, 'G': 13, 'H': 10, 'I': 18, 'J': 13, 'K': 38, 'L': 23, 
        'M': 59, 'N': 19, 'O': 18, 'P': 18, 'Q': 50, 'R': 18, 'S': 44
    }
    set_column_widths(worksheet, column_widths)
    
    # Color the Simple COA column header light blue
    worksheet['N1'].fill = light_blue_fill  # Simple COA column

    # Families sheet formatting
    worksheet = workbook['Families']
    filters = worksheet.auto_filter
    filters.ref = 'A1:E' + str(worksheet.max_row)
    column_widths = {
        'A': 14, 'B': 38, 'C': 109, 'D': 15, 'E': 15
    }
    set_column_widths(worksheet, column_widths)
    
    # Color the Name(s) and Full Email(s) column headers light blue
    worksheet['B1'].fill = light_blue_fill  # Name(s) column
    worksheet['C1'].fill = light_blue_fill  # Full Email(s) column

    workbook.save(output_filename)
    typer.echo(f'Wrote ILW data to {output_filename}')

    # Output recharacterization dictionary only if logging level is DEBUG
    if logging.getLogger().getEffectiveLevel() <= logging.DEBUG:
        typer.echo("\n" + "="*80)
        typer.echo("RECHARACTERIZATIONS SUMMARY")
        typer.echo("="*80)
        if recharacterizations:
            typer.echo(f"\nFound {len(recharacterizations)} recharacterization(s):\n")
            total_amount = 0
            for find_code, amount in sorted(recharacterizations.items()):
                typer.echo(f"  {find_code}: ${amount:,.2f}")
                total_amount += amount
            typer.echo(f"\nTotal amount to be recharacterized to Projects: ${total_amount:,.2f}")
        else:
            typer.echo("\nNo recharacterizations found.")
        typer.echo("="*80 + "\n")

if __name__ == "__main__":
    app() 
